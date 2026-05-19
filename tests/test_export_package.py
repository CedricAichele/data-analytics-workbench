from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from app.services.export_package import build_export_workbook, safe_excel_sheet_name
from app.services.quality_score import calculate_quality_score


def test_safe_excel_sheet_name_removes_invalid_characters_and_limits_length():
    name = safe_excel_sheet_name("Very/Long:Invalid*Sheet?Name[That]ExceedsLimit")

    assert "/" not in name
    assert ":" not in name
    assert len(name) <= 31


def test_build_export_workbook_creates_expected_sheets():
    cleaned = pd.DataFrame({"a": [1, 2], "b": ["x", "y"]})
    dictionary = pd.DataFrame({"column_name": ["a", "b"], "detected_data_type": ["numeric", "text"]})
    quality_rules = pd.DataFrame(
        {
            "template": ["Sales / Retail"],
            "rule_name": ["unit_price <= 0"],
            "severity": ["critical"],
            "affected_rows_count": [1],
            "affected_percentage": [50.0],
            "explanation": ["Invalid price"],
            "recommended_fix": ["Correct price"],
        }
    )

    workbook_bytes = build_export_workbook(
        cleaned_data=cleaned,
        data_dictionary=dictionary,
        quality_report=calculate_quality_score(cleaned),
        quality_rules=quality_rules,
        transformation_log=["Rename column: old -> new"],
        generic_analytics_result=pd.DataFrame({"metric": ["sum"], "a": [3]}),
        kpi_summary=pd.DataFrame({"source": ["Sales"], "metric": ["gross_revenue"], "value": [100.0]}),
        result_tables={
            "Sales_Results": pd.DataFrame({"metric": ["gross_revenue"], "value": [100.0]}),
            "Result_Tables": pd.DataFrame({"source_table": ["Sales chart"], "value": [100.0]}),
        },
    )

    assert len(workbook_bytes) > 0
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)
    assert {
        "Cleaned_Data",
        "Data_Dictionary",
        "Data_Quality",
        "Transformation_Log",
        "KPI_Summary",
        "Quality_Rules",
        "Generic_Analytics_Result",
        "Sales_Results",
        "Result_Tables",
    }.issubset(set(workbook.sheetnames))

    kpi_header = [cell.value for cell in next(workbook["KPI_Summary"].iter_rows(max_row=1))]
    assert kpi_header == ["source", "metric", "value"]
